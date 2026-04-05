"""
Report Export Service - Generate Detailed Excel, PDF, and XML reports (Checkmarx-style)
Includes: SAST, SCA, Secrets Detection, Threat Model (STRIDE), and MITRE ATT&CK Mapping
"""
from typing import List, Dict, Any, Tuple
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from io import BytesIO

class ReportService:
    """Service for generating detailed security reports in Checkmarx style"""

    # Severity colors
    SEVERITY_COLORS = {
        'critical': 'C00000',  # Dark Red
        'high': 'FF0000',      # Red
        'medium': 'FFA500',    # Orange
        'low': 'FFFF00',       # Yellow
        'info': '92D050'       # Green
    }

    SEVERITY_ORDER = ['critical', 'high', 'medium', 'low', 'info']

    def generate_excel_report(self, scan_data: Dict[str, Any], output_path: str = None) -> BytesIO:
        """
        Generate comprehensive Excel report with multiple detailed sheets (Checkmarx-style)

        Sheets:
        1. Cover Page - Project metadata and branding
        2. Dashboard - Visual summary with statistics
        3. Overall Summary - Aggregated findings across all scan types
        4-7. SAST Findings by Severity (Critical, High, Medium, Low)
        8. SAST - Grouped by CWE
        9. SAST - Grouped by OWASP Category
        10. SAST - Grouped by File
        11-14. SCA Findings by Severity (Critical, High, Medium, Low)
        15. SCA - Grouped by Package
        16. SCA - CVE Details
        17. Secrets Detection - All findings
        18. Threat Model - STRIDE Analysis
        19. MITRE ATT&CK Mapping
        20. Remediation Summary - Consolidated recommendations
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Common styles
        styles = self._create_styles()

        # Sheet 1: Cover Page
        self._create_cover_page(wb, scan_data, styles)

        # Sheet 2: Dashboard
        self._create_dashboard_sheet(wb, scan_data, styles)

        # Sheet 3: Overall Summary
        self._create_overall_summary_sheet(wb, scan_data, styles)

        # SAST Findings - Detailed sheets by severity
        sast_findings = scan_data.get('sast_findings', [])
        if sast_findings:
            for severity in self.SEVERITY_ORDER:
                findings = [f for f in sast_findings if f.get('severity', '').lower() == severity]
                if findings:
                    self._create_sast_detailed_sheet(wb, findings, severity, styles)

            # SAST - Grouped views
            self._create_sast_by_cwe_sheet(wb, sast_findings, styles)
            self._create_sast_by_owasp_sheet(wb, sast_findings, styles)
            self._create_sast_by_file_sheet(wb, sast_findings, styles)

        # SCA Findings - Detailed sheets by severity
        sca_findings = scan_data.get('sca_findings', [])
        if sca_findings:
            for severity in self.SEVERITY_ORDER:
                findings = [f for f in sca_findings if f.get('severity', '').lower() == severity]
                if findings:
                    self._create_sca_detailed_sheet(wb, findings, severity, styles)

            # SCA - Grouped views
            self._create_sca_by_package_sheet(wb, sca_findings, styles)
            self._create_sca_cve_details_sheet(wb, sca_findings, styles)

        # Secrets Detection
        secret_findings = scan_data.get('secret_findings', [])
        if secret_findings:
            self._create_secrets_detailed_sheet(wb, secret_findings, styles)

        # Threat Model (STRIDE)
        threat_model = scan_data.get('threat_model', {})
        if threat_model:
            self._create_threat_model_sheet(wb, threat_model, styles)
            self._create_mitre_attack_sheet(wb, threat_model, styles)

        # GitHub Monitor
        github_monitor = scan_data.get('github_monitor', {})
        if github_monitor:
            self._create_github_monitor_sheet(wb, github_monitor, styles)

        # Remediation Summary
        self._create_remediation_summary_sheet(wb, scan_data, styles)

        # Save or return
        if output_path:
            wb.save(output_path)
            return output_path
        else:
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

    def _create_styles(self) -> Dict[str, Any]:
        """Create common styles for Excel sheets"""
        return {
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_font': Font(color="FFFFFF", bold=True, size=12),
            'title_font': Font(bold=True, size=20, color="366092"),
            'subtitle_font': Font(bold=True, size=14, color="366092"),
            'bold_font': Font(bold=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left_align': Alignment(horizontal='left', vertical='top', wrap_text=True),
        }

    def _create_cover_page(self, wb: Workbook, scan_data: Dict[str, Any], styles: Dict):
        """Create cover page with project metadata"""
        ws = wb.create_sheet("Cover Page")

        # Title
        ws['B2'] = "Application Security Scan Report"
        ws['B2'].font = Font(bold=True, size=24, color="366092")
        ws.merge_cells('B2:G2')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        # Subtitle
        ws['B4'] = "Detailed Security Analysis Report"
        ws['B4'].font = Font(size=14, color="666666")
        ws.merge_cells('B4:G4')
        ws['B4'].alignment = Alignment(horizontal='center')

        # Project Information
        row = 7
        metadata = [
            ('Project Name:', scan_data.get('project_name', 'Unknown')),
            ('Report Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('Scan Types:', ', '.join(scan_data.get('scan_types', []))),
            ('Total Findings:', str(self._get_total_findings(scan_data))),
        ]

        for label, value in metadata:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = styles['bold_font']
            ws[f'D{row}'] = value
            row += 1

        # Severity Summary
        row += 2
        ws[f'B{row}'] = "Severity Distribution"
        ws[f'B{row}'].font = styles['subtitle_font']
        row += 1

        severity_counts = self._count_all_severities(scan_data)
        for severity in self.SEVERITY_ORDER:
            count = severity_counts.get(severity, 0)
            if count > 0:
                ws[f'B{row}'] = severity.upper()
                ws[f'D{row}'] = count
                ws[f'B{row}'].fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                  end_color=self.SEVERITY_COLORS[severity],
                                                  fill_type="solid")
                if severity in ['critical', 'high']:
                    ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
                row += 1

        # Column widths
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 30

    def _create_dashboard_sheet(self, wb: Workbook, scan_data: Dict[str, Any], styles: Dict):
        """Create dashboard with summary statistics"""
        ws = wb.create_sheet("Dashboard")

        # Title
        ws['A1'] = "Security Scan Dashboard"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Overall Statistics
        row = 3
        ws[f'A{row}'] = "Overall Statistics"
        ws[f'A{row}'].font = styles['subtitle_font']
        row += 1

        # Headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        # Metrics
        sast_findings = scan_data.get('sast_findings', [])
        sca_findings = scan_data.get('sca_findings', [])
        secret_findings = scan_data.get('secret_findings', [])

        metrics = [
            ('Total Vulnerabilities', self._get_total_findings(scan_data)),
            ('SAST Findings', len(sast_findings)),
            ('SCA Findings', len(sca_findings)),
            ('Secrets Detected', len(secret_findings)),
            ('Files Scanned', len(set([f.get('file_path', '') for f in sast_findings + secret_findings]))),
            ('Unique CWEs', len(set([f.get('cwe_id', '') for f in sast_findings if f.get('cwe_id')]))),
            ('Unique CVEs', len(set([f.get('cve', '') for f in sca_findings if f.get('cve')]))),
        ]

        for metric, value in metrics:
            ws.cell(row, 1, metric)
            ws.cell(row, 2, value)
            row += 1

        # Severity Distribution Table
        row += 2
        ws[f'A{row}'] = "Severity Distribution"
        ws[f'A{row}'].font = styles['subtitle_font']
        row += 1

        headers = ['Severity', 'SAST', 'SCA', 'Secrets', 'Total', 'Percentage']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        sast_severity = self._count_by_severity(sast_findings)
        sca_severity = self._count_by_severity(sca_findings)
        secret_severity = self._count_by_severity(secret_findings)
        total_findings = self._get_total_findings(scan_data)

        for severity in self.SEVERITY_ORDER:
            sast_count = sast_severity.get(severity, 0)
            sca_count = sca_severity.get(severity, 0)
            secret_count = secret_severity.get(severity, 0)
            total = sast_count + sca_count + secret_count
            percentage = (total / total_findings * 100) if total_findings > 0 else 0

            ws.cell(row, 1, severity.upper())
            ws.cell(row, 2, sast_count)
            ws.cell(row, 3, sca_count)
            ws.cell(row, 4, secret_count)
            ws.cell(row, 5, total)
            ws.cell(row, 6, f"{percentage:.1f}%")

            # Color code severity column
            severity_cell = ws.cell(row, 1)
            severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                            end_color=self.SEVERITY_COLORS[severity],
                                            fill_type="solid")
            if severity in ['critical', 'high']:
                severity_cell.font = Font(color="FFFFFF", bold=True)

            row += 1

        # Auto-size columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_overall_summary_sheet(self, wb: Workbook, scan_data: Dict[str, Any], styles: Dict):
        """Create overall summary sheet with aggregated findings"""
        ws = wb.create_sheet("Overall Summary")

        # Title
        ws['A1'] = "Overall Security Summary"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:J1')

        # Scan Information
        row = 3
        scan_info = [
            ('Project:', scan_data.get('project_name', 'Unknown')),
            ('Scan Date:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('Scan Types:', ', '.join(scan_data.get('scan_types', []))),
        ]

        for label, value in scan_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = styles['bold_font']
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            row += 1

        # Summary Table
        row += 2
        ws[f'A{row}'] = "Findings Summary by Scan Type and Severity"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:J{row}')
        row += 1

        headers = ['Scan Type', 'Total', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Risk Score']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']

        row += 1

        # Data rows
        scan_types_data = [
            ('SAST (Code Vulnerabilities)', scan_data.get('sast_findings', [])),
            ('SCA (Dependencies)', scan_data.get('sca_findings', [])),
            ('Secrets Detection', scan_data.get('secret_findings', [])),
        ]

        for scan_type, findings in scan_types_data:
            if findings:
                severity_counts = self._count_by_severity(findings)
                risk_score = self._calculate_risk_score(severity_counts)

                ws.cell(row, 1, scan_type)
                ws.cell(row, 2, len(findings))
                ws.cell(row, 3, severity_counts.get('critical', 0))
                ws.cell(row, 4, severity_counts.get('high', 0))
                ws.cell(row, 5, severity_counts.get('medium', 0))
                ws.cell(row, 6, severity_counts.get('low', 0))
                ws.cell(row, 7, severity_counts.get('info', 0))
                ws.cell(row, 8, f"{risk_score:.1f}")

                row += 1

        # Total row
        total_severity = self._count_all_severities(scan_data)
        total_risk = self._calculate_risk_score(total_severity)

        ws.cell(row, 1, 'TOTAL')
        ws.cell(row, 1).font = styles['bold_font']
        ws.cell(row, 2, self._get_total_findings(scan_data))
        ws.cell(row, 3, total_severity.get('critical', 0))
        ws.cell(row, 4, total_severity.get('high', 0))
        ws.cell(row, 5, total_severity.get('medium', 0))
        ws.cell(row, 6, total_severity.get('low', 0))
        ws.cell(row, 7, total_severity.get('info', 0))
        ws.cell(row, 8, f"{total_risk:.1f}")

        # Bold total row
        for col in range(1, 9):
            ws.cell(row, col).font = styles['bold_font']

        # Top Issues Section
        row += 3
        ws[f'A{row}'] = "Top 10 Critical Issues Requiring Immediate Attention"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:J{row}')
        row += 1

        headers = ['#', 'Type', 'Severity', 'Title', 'File/Package', 'Line/Version', 'CWE/CVE', 'CVSS']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        # Get top critical/high findings
        all_findings = []
        for finding in scan_data.get('sast_findings', []):
            all_findings.append({**finding, 'type': 'SAST'})
        for finding in scan_data.get('sca_findings', []):
            all_findings.append({**finding, 'type': 'SCA'})
        for finding in scan_data.get('secret_findings', []):
            all_findings.append({**finding, 'type': 'SECRET'})

        # Sort by severity
        severity_priority = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}
        all_findings.sort(key=lambda x: (severity_priority.get(x.get('severity', 'info').lower(), 5),
                                         -float(x.get('cvss_score', 0))))

        for idx, finding in enumerate(all_findings[:10], start=1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, finding['type'])
            ws.cell(row, 3, finding.get('severity', '').upper())
            ws.cell(row, 4, finding.get('title', finding.get('vulnerability', finding.get('secret_type', 'N/A'))))
            ws.cell(row, 5, finding.get('file_path', finding.get('package', 'N/A')))
            ws.cell(row, 6, finding.get('line_number', finding.get('installed_version', 'N/A')))
            ws.cell(row, 7, finding.get('cwe_id', finding.get('cve', 'N/A')))
            ws.cell(row, 8, finding.get('cvss_score', 'N/A'))

            # Color code severity
            severity = finding.get('severity', '').lower()
            if severity in self.SEVERITY_COLORS:
                severity_cell = ws.cell(row, 3)
                severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                end_color=self.SEVERITY_COLORS[severity],
                                                fill_type="solid")
                if severity in ['critical', 'high']:
                    severity_cell.font = Font(color="FFFFFF", bold=True)

            row += 1

        # Auto-size columns
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_sast_detailed_sheet(self, wb: Workbook, findings: List[Dict], severity: str, styles: Dict):
        """Create detailed SAST findings sheet with analyst workflow columns"""
        from openpyxl.worksheet.datavalidation import DataValidation

        sheet_name = f"SAST - {severity.upper()}"
        ws = wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f"SAST Findings - {severity.upper()} Severity"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:T1')

        # Count & open/resolved stats
        open_count = sum(1 for f in findings if f.get('status', 'Open') == 'Open')
        ws['A2'] = f"Total: {len(findings)} | Open: {open_count} | Resolved: {len(findings) - open_count}"
        ws['A2'].font = styles['bold_font']

        # Headers - enhanced with analyst workflow columns
        row = 4
        headers = ['#', 'Title', 'Severity', 'CWE ID', 'OWASP', 'File Path', 'Line', 'CVSS',
                   'STRIDE Category', 'MITRE ATT&CK', 'Description', 'Business Impact', 'Technical Impact',
                   'Remediation', 'Code Snippet',
                   'Status', 'Assignee', 'Due Date', 'Priority', 'Notes']

        # Color-code analyst columns differently
        analyst_col_start = 16  # Status column index
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            if col >= analyst_col_start:
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            else:
                cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']

        row += 1

        # Data validation for analyst columns
        status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,False Positive,Accepted Risk,Deferred"', allow_blank=True)
        status_dv.prompt = "Select finding status"
        priority_dv = DataValidation(type="list", formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low,P5 - Info"', allow_blank=True)
        ws.add_data_validation(status_dv)
        ws.add_data_validation(priority_dv)

        # SLA mapping for due dates
        sla_days = {'critical': 7, 'high': 14, 'medium': 30, 'low': 90, 'info': 180}
        from datetime import timedelta

        for idx, finding in enumerate(findings, start=1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, finding.get('title', 'N/A'))
            ws.cell(row, 3, finding.get('severity', '').upper())
            ws.cell(row, 4, finding.get('cwe_id', 'N/A'))
            ws.cell(row, 5, finding.get('owasp_category', 'N/A'))
            ws.cell(row, 6, finding.get('file_path', 'N/A'))
            ws.cell(row, 7, finding.get('line_number', 'N/A'))
            ws.cell(row, 8, finding.get('cvss_score', 'N/A'))
            ws.cell(row, 9, finding.get('stride_category', 'N/A'))
            ws.cell(row, 10, f"{finding.get('mitre_attack_id', '')} {finding.get('mitre_attack_name', '')}".strip() or 'N/A')
            ws.cell(row, 11, finding.get('description', 'N/A'))
            ws.cell(row, 12, finding.get('business_impact', 'N/A'))
            ws.cell(row, 13, finding.get('technical_impact', 'N/A'))
            ws.cell(row, 14, finding.get('remediation', 'N/A'))
            ws.cell(row, 15, finding.get('code_snippet', 'N/A'))
            # Analyst workflow columns
            ws.cell(row, 16, finding.get('status', 'Open'))
            ws.cell(row, 17, '')  # Assignee - blank for analyst to fill
            sev = finding.get('severity', 'info').lower()
            due = datetime.now() + timedelta(days=sla_days.get(sev, 90))
            ws.cell(row, 18, due.strftime('%Y-%m-%d'))
            ws.cell(row, 19, f"P{['critical','high','medium','low','info'].index(sev)+1} - {sev.capitalize()}" if sev in sla_days else 'P4 - Low')
            ws.cell(row, 20, '')  # Notes - blank for analyst

            # Add data validations
            status_dv.add(ws.cell(row, 16))
            priority_dv.add(ws.cell(row, 19))

            # Color code severity
            severity_cell = ws.cell(row, 3)
            severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity.lower()],
                                            end_color=self.SEVERITY_COLORS[severity.lower()],
                                            fill_type="solid")
            if severity.lower() in ['critical', 'high']:
                severity_cell.font = Font(color="FFFFFF", bold=True)

            # Color code status
            status_val = finding.get('status', 'Open')
            status_colors = {'Open': 'FF0000', 'In Progress': 'FFA500', 'Resolved': '92D050', 'False Positive': '808080', 'Accepted Risk': '4472C4'}
            if status_val in status_colors:
                ws.cell(row, 16).fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
                if status_val in ['Open']:
                    ws.cell(row, 16).font = Font(color="FFFFFF", bold=True)

            # Light green background for analyst columns (skip status cell if already colored)
            for c in range(analyst_col_start, 21):
                if c == 16 and status_val in status_colors:
                    continue
                ws.cell(row, c).fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

            ws.row_dimensions[row].height = 60
            for c in range(1, 21):
                ws.cell(row, c).border = styles['border']
                ws.cell(row, c).alignment = styles['left_align']

            row += 1

        # Auto-size columns
        column_widths = [5, 30, 12, 12, 20, 35, 8, 8, 18, 25, 40, 30, 30, 40, 40, 15, 15, 12, 15, 30]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A5'

    def _create_sast_by_cwe_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create SAST findings grouped by CWE"""
        ws = wb.create_sheet("SAST - By CWE")

        # Title
        ws['A1'] = "SAST Findings Grouped by CWE"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Group by CWE
        cwe_groups = defaultdict(list)
        for finding in findings:
            cwe = finding.get('cwe_id', 'Unknown')
            cwe_groups[cwe].append(finding)

        # Sort by count
        sorted_cwes = sorted(cwe_groups.items(), key=lambda x: len(x[1]), reverse=True)

        row = 3
        for cwe_id, cwe_findings in sorted_cwes:
            # CWE Header
            cwe_name = cwe_findings[0].get('cwe_name', 'Unknown CWE')
            ws[f'A{row}'] = f"{cwe_id}: {cwe_name}"
            ws[f'A{row}'].font = styles['subtitle_font']
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            ws[f'A{row}'] = f"Total Findings: {len(cwe_findings)}"
            ws[f'A{row}'].font = styles['bold_font']
            row += 1

            # Headers
            headers = ['#', 'Severity', 'Title', 'File Path', 'Line', 'CVSS', 'OWASP', 'Remediation']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']

            row += 1

            # Findings
            for idx, finding in enumerate(cwe_findings, start=1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, finding.get('severity', '').upper())
                ws.cell(row, 3, finding.get('title', 'N/A'))
                ws.cell(row, 4, finding.get('file_path', 'N/A'))
                ws.cell(row, 5, finding.get('line_number', 'N/A'))
                ws.cell(row, 6, finding.get('cvss_score', 'N/A'))
                ws.cell(row, 7, finding.get('owasp_category', 'N/A'))
                ws.cell(row, 8, finding.get('remediation', 'N/A'))

                # Color code severity
                severity = finding.get('severity', '').lower()
                if severity in self.SEVERITY_COLORS:
                    severity_cell = ws.cell(row, 2)
                    severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                    end_color=self.SEVERITY_COLORS[severity],
                                                    fill_type="solid")
                    if severity in ['critical', 'high']:
                        severity_cell.font = Font(color="FFFFFF", bold=True)

                row += 1

            row += 2  # Space between groups

        # Auto-size columns
        column_widths = [5, 12, 35, 40, 8, 8, 20, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_sast_by_owasp_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create SAST findings grouped by OWASP category"""
        ws = wb.create_sheet("SAST - By OWASP")

        # Title
        ws['A1'] = "SAST Findings Grouped by OWASP Top 10"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Group by OWASP
        owasp_groups = defaultdict(list)
        for finding in findings:
            owasp = finding.get('owasp_category', 'Uncategorized')
            owasp_groups[owasp].append(finding)

        # Sort by OWASP category
        sorted_owasp = sorted(owasp_groups.items(), key=lambda x: x[0])

        row = 3
        for owasp_cat, owasp_findings in sorted_owasp:
            # OWASP Header
            ws[f'A{row}'] = owasp_cat
            ws[f'A{row}'].font = styles['subtitle_font']
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            ws[f'A{row}'] = f"Total Findings: {len(owasp_findings)}"
            ws[f'A{row}'].font = styles['bold_font']
            row += 1

            # Headers
            headers = ['#', 'Severity', 'Title', 'CWE', 'File Path', 'Line', 'CVSS', 'Remediation']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']

            row += 1

            # Findings
            for idx, finding in enumerate(owasp_findings, start=1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, finding.get('severity', '').upper())
                ws.cell(row, 3, finding.get('title', 'N/A'))
                ws.cell(row, 4, finding.get('cwe_id', 'N/A'))
                ws.cell(row, 5, finding.get('file_path', 'N/A'))
                ws.cell(row, 6, finding.get('line_number', 'N/A'))
                ws.cell(row, 7, finding.get('cvss_score', 'N/A'))
                ws.cell(row, 8, finding.get('remediation', 'N/A'))

                # Color code severity
                severity = finding.get('severity', '').lower()
                if severity in self.SEVERITY_COLORS:
                    severity_cell = ws.cell(row, 2)
                    severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                    end_color=self.SEVERITY_COLORS[severity],
                                                    fill_type="solid")
                    if severity in ['critical', 'high']:
                        severity_cell.font = Font(color="FFFFFF", bold=True)

                row += 1

            row += 2

        # Auto-size columns
        column_widths = [5, 12, 35, 15, 40, 8, 8, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_sast_by_file_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create SAST findings grouped by file"""
        ws = wb.create_sheet("SAST - By File")

        # Title
        ws['A1'] = "SAST Findings Grouped by File"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Group by file
        file_groups = defaultdict(list)
        for finding in findings:
            file_path = finding.get('file_path', 'Unknown')
            file_groups[file_path].append(finding)

        # Sort by count
        sorted_files = sorted(file_groups.items(), key=lambda x: len(x[1]), reverse=True)

        row = 3
        for file_path, file_findings in sorted_files:
            # File Header
            ws[f'A{row}'] = file_path
            ws[f'A{row}'].font = styles['subtitle_font']
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            ws[f'A{row}'] = f"Total Findings: {len(file_findings)}"
            ws[f'A{row}'].font = styles['bold_font']
            row += 1

            # Headers
            headers = ['#', 'Severity', 'Title', 'CWE', 'Line', 'CVSS', 'OWASP', 'Remediation']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']

            row += 1

            # Sort findings by line number
            file_findings.sort(key=lambda x: x.get('line_number', 0))

            # Findings
            for idx, finding in enumerate(file_findings, start=1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, finding.get('severity', '').upper())
                ws.cell(row, 3, finding.get('title', 'N/A'))
                ws.cell(row, 4, finding.get('cwe_id', 'N/A'))
                ws.cell(row, 5, finding.get('line_number', 'N/A'))
                ws.cell(row, 6, finding.get('cvss_score', 'N/A'))
                ws.cell(row, 7, finding.get('owasp_category', 'N/A'))
                ws.cell(row, 8, finding.get('remediation', 'N/A'))

                # Color code severity
                severity = finding.get('severity', '').lower()
                if severity in self.SEVERITY_COLORS:
                    severity_cell = ws.cell(row, 2)
                    severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                    end_color=self.SEVERITY_COLORS[severity],
                                                    fill_type="solid")
                    if severity in ['critical', 'high']:
                        severity_cell.font = Font(color="FFFFFF", bold=True)

                row += 1

            row += 2

        # Auto-size columns
        column_widths = [5, 12, 35, 15, 8, 8, 20, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_sca_detailed_sheet(self, wb: Workbook, findings: List[Dict], severity: str, styles: Dict):
        """Create detailed SCA findings sheet with reachability and analyst workflow columns"""
        from openpyxl.worksheet.datavalidation import DataValidation
        from datetime import timedelta

        sheet_name = f"SCA - {severity.upper()}"
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"SCA (Dependency) Findings - {severity.upper()} Severity"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:S1')

        open_count = sum(1 for f in findings if f.get('status', 'Open') == 'Open')
        ws['A2'] = f"Total: {len(findings)} | Open: {open_count} | Resolved: {len(findings) - open_count}"
        ws['A2'].font = styles['bold_font']

        row = 4
        headers = ['#', 'Package Name', 'Installed Version', 'Fixed Version', 'Vulnerability',
                   'CVE', 'Severity', 'CVSS', 'CWE ID', 'Description', 'Remediation',
                   'Reachability', 'Business Impact', 'Technical Impact',
                   'Status', 'Assignee', 'Due Date', 'Priority', 'Notes']

        analyst_col_start = 15
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx, header)
            if col_idx >= analyst_col_start:
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            elif col_idx == 12:  # Reachability column
                cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
            else:
                cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']

        row += 1

        status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,False Positive,Accepted Risk,Deferred"', allow_blank=True)
        priority_dv = DataValidation(type="list", formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low,P5 - Info"', allow_blank=True)
        reachability_dv = DataValidation(type="list", formula1='"Exploitable,Potentially Exploitable,Imported Only,Not Reachable,Not Analyzed"', allow_blank=True)
        ws.add_data_validation(status_dv)
        ws.add_data_validation(priority_dv)
        ws.add_data_validation(reachability_dv)

        sla_days = {'critical': 7, 'high': 14, 'medium': 30, 'low': 90, 'info': 180}

        for idx, finding in enumerate(findings, start=1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, finding.get('package', 'N/A'))
            ws.cell(row, 3, finding.get('installed_version', 'N/A'))
            ws.cell(row, 4, finding.get('fixed_version', finding.get('safe_version', 'N/A')))
            ws.cell(row, 5, finding.get('vulnerability', finding.get('title', 'N/A')))
            ws.cell(row, 6, finding.get('cve', 'N/A'))
            ws.cell(row, 7, finding.get('severity', '').upper())
            ws.cell(row, 8, finding.get('cvss_score', 'N/A'))
            ws.cell(row, 9, finding.get('cwe_id', 'N/A'))
            ws.cell(row, 10, finding.get('description', 'N/A'))
            ws.cell(row, 11, finding.get('remediation', 'N/A'))
            # Reachability - default to Not Analyzed
            ws.cell(row, 12, finding.get('reachability', 'Not Analyzed'))
            ws.cell(row, 13, finding.get('business_impact', 'N/A'))
            ws.cell(row, 14, finding.get('technical_impact', 'N/A'))
            # Analyst workflow
            ws.cell(row, 15, finding.get('status', 'Open'))
            ws.cell(row, 16, '')
            sev = finding.get('severity', 'info').lower()
            due = datetime.now() + timedelta(days=sla_days.get(sev, 90))
            ws.cell(row, 17, due.strftime('%Y-%m-%d'))
            ws.cell(row, 18, f"P{['critical','high','medium','low','info'].index(sev)+1} - {sev.capitalize()}" if sev in sla_days else 'P4 - Low')
            ws.cell(row, 19, '')

            status_dv.add(ws.cell(row, 15))
            priority_dv.add(ws.cell(row, 18))
            reachability_dv.add(ws.cell(row, 12))

            # Color code severity
            severity_cell = ws.cell(row, 7)
            severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity.lower()],
                                            end_color=self.SEVERITY_COLORS[severity.lower()],
                                            fill_type="solid")
            if severity.lower() in ['critical', 'high']:
                severity_cell.font = Font(color="FFFFFF", bold=True)

            # Color code reachability
            reach_colors = {'Exploitable': 'C00000', 'Potentially Exploitable': 'FFA500', 'Imported Only': 'FFFF00', 'Not Reachable': '92D050', 'Not Analyzed': 'D9D9D9'}
            reach_val = finding.get('reachability', 'Not Analyzed')
            if reach_val in reach_colors:
                ws.cell(row, 12).fill = PatternFill(start_color=reach_colors[reach_val], end_color=reach_colors[reach_val], fill_type="solid")
                if reach_val in ['Exploitable']:
                    ws.cell(row, 12).font = Font(color="FFFFFF", bold=True)

            # Color code status
            status_val = finding.get('status', 'Open')
            status_colors = {'Open': 'FF0000', 'In Progress': 'FFA500', 'Resolved': '92D050', 'False Positive': '808080', 'Accepted Risk': '4472C4'}
            if status_val in status_colors:
                ws.cell(row, 15).fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
                if status_val == 'Open':
                    ws.cell(row, 15).font = Font(color="FFFFFF", bold=True)

            ws.row_dimensions[row].height = 60
            for c in range(1, 20):
                ws.cell(row, c).border = styles['border']
                ws.cell(row, c).alignment = styles['left_align']

            row += 1

        column_widths = [5, 30, 15, 15, 30, 18, 12, 8, 12, 40, 40, 20, 30, 30, 15, 15, 12, 15, 30]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = 'A5'

    def _create_sca_by_package_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create SCA findings grouped by package"""
        ws = wb.create_sheet("SCA - By Package")

        # Title
        ws['A1'] = "SCA Findings Grouped by Package"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Group by package
        package_groups = defaultdict(list)
        for finding in findings:
            package = finding.get('package', 'Unknown')
            package_groups[package].append(finding)

        # Sort by package name
        sorted_packages = sorted(package_groups.items())

        row = 3
        for package_name, package_findings in sorted_packages:
            # Package Header
            ws[f'A{row}'] = package_name
            ws[f'A{row}'].font = styles['subtitle_font']
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            ws[f'A{row}'] = f"Total Vulnerabilities: {len(package_findings)} | Installed Version: {package_findings[0].get('installed_version', 'N/A')}"
            ws[f'A{row}'].font = styles['bold_font']
            row += 1

            # Headers
            headers = ['#', 'Severity', 'CVE', 'Vulnerability', 'CVSS', 'Fixed Version', 'Description', 'Remediation']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']

            row += 1

            # Findings
            for idx, finding in enumerate(package_findings, start=1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, finding.get('severity', '').upper())
                ws.cell(row, 3, finding.get('cve', 'N/A'))
                ws.cell(row, 4, finding.get('vulnerability', 'N/A'))
                ws.cell(row, 5, finding.get('cvss_score', 'N/A'))
                ws.cell(row, 6, finding.get('fixed_version', finding.get('safe_version', 'N/A')))
                ws.cell(row, 7, finding.get('description', 'N/A'))
                ws.cell(row, 8, finding.get('remediation', 'N/A'))

                # Color code severity
                severity = finding.get('severity', '').lower()
                if severity in self.SEVERITY_COLORS:
                    severity_cell = ws.cell(row, 2)
                    severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                    end_color=self.SEVERITY_COLORS[severity],
                                                    fill_type="solid")
                    if severity in ['critical', 'high']:
                        severity_cell.font = Font(color="FFFFFF", bold=True)

                row += 1

            row += 2

        # Auto-size columns
        column_widths = [5, 12, 18, 30, 8, 15, 40, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_sca_cve_details_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create detailed CVE information sheet"""
        ws = wb.create_sheet("SCA - CVE Details")

        # Title
        ws['A1'] = "Detailed CVE Information"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:J1')

        # Headers
        row = 3
        headers = ['CVE ID', 'Severity', 'CVSS Score', 'Package', 'Version', 'Description',
                   'References', 'Published Date', 'Remediation', 'Status']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        # Get unique CVEs
        cve_dict = {}
        for finding in findings:
            cve = finding.get('cve', '')
            if cve and cve not in cve_dict:
                cve_dict[cve] = finding

        # Sort by CVSS score
        sorted_cves = sorted(cve_dict.values(),
                           key=lambda x: float(x.get('cvss_score', 0)),
                           reverse=True)

        for finding in sorted_cves:
            ws.cell(row, 1, finding.get('cve', 'N/A'))
            ws.cell(row, 2, finding.get('severity', '').upper())
            ws.cell(row, 3, finding.get('cvss_score', 'N/A'))
            ws.cell(row, 4, finding.get('package', 'N/A'))
            ws.cell(row, 5, finding.get('installed_version', 'N/A'))
            ws.cell(row, 6, finding.get('description', 'N/A'))
            ws.cell(row, 7, finding.get('references', 'N/A'))
            ws.cell(row, 8, finding.get('published_date', 'N/A'))
            ws.cell(row, 9, finding.get('remediation', 'N/A'))
            ws.cell(row, 10, 'Open' if finding.get('severity', '').lower() in ['critical', 'high'] else 'Review')

            # Color code severity
            severity = finding.get('severity', '').lower()
            if severity in self.SEVERITY_COLORS:
                severity_cell = ws.cell(row, 2)
                severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                end_color=self.SEVERITY_COLORS[severity],
                                                fill_type="solid")
                if severity in ['critical', 'high']:
                    severity_cell.font = Font(color="FFFFFF", bold=True)

            ws.row_dimensions[row].height = 50

            for col in range(1, 11):
                ws.cell(row, col).border = styles['border']
                ws.cell(row, col).alignment = styles['left_align']

            row += 1

        # Auto-size columns
        column_widths = [18, 12, 10, 25, 15, 40, 30, 15, 40, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_secrets_detailed_sheet(self, wb: Workbook, findings: List[Dict], styles: Dict):
        """Create detailed secrets detection sheet with analyst workflow columns"""
        from openpyxl.worksheet.datavalidation import DataValidation
        from datetime import timedelta

        ws = wb.create_sheet("Secrets Detection")

        ws['A1'] = "Secrets Detection - All Findings"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:N1')

        open_count = sum(1 for f in findings if f.get('status', 'Open') == 'Open')
        ws['A2'] = f"Total Secrets: {len(findings)} | Open: {open_count} | Resolved: {len(findings) - open_count}"
        ws['A2'].font = Font(bold=True, size=12, color="C00000")

        row = 4
        headers = ['#', 'Secret Type', 'Severity', 'CVSS', 'CWE ID', 'File Path', 'Line',
                   'Description', 'Remediation',
                   'Status', 'Assignee', 'Due Date', 'Priority', 'Notes']

        analyst_col_start = 10
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx, header)
            if col_idx >= analyst_col_start:
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            else:
                cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']

        row += 1

        status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,False Positive,Accepted Risk,Deferred"', allow_blank=True)
        priority_dv = DataValidation(type="list", formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low"', allow_blank=True)
        ws.add_data_validation(status_dv)
        ws.add_data_validation(priority_dv)

        severity_priority = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}
        findings.sort(key=lambda x: severity_priority.get(x.get('severity', 'info').lower(), 5))

        sla_days = {'critical': 3, 'high': 7, 'medium': 14, 'low': 30, 'info': 90}

        for idx, finding in enumerate(findings, start=1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, finding.get('secret_type', finding.get('title', 'Unknown')))
            ws.cell(row, 3, finding.get('severity', '').upper())
            ws.cell(row, 4, finding.get('cvss_score', 'N/A'))
            ws.cell(row, 5, finding.get('cwe_id', 'CWE-798'))
            ws.cell(row, 6, finding.get('file_path', 'N/A'))
            ws.cell(row, 7, finding.get('line_number', 'N/A'))
            ws.cell(row, 8, finding.get('description', 'Hardcoded secret detected'))
            ws.cell(row, 9, finding.get('remediation', 'Remove hardcoded secret and use environment variables'))
            # Analyst workflow
            ws.cell(row, 10, finding.get('status', 'Open'))
            ws.cell(row, 11, '')
            sev = finding.get('severity', 'high').lower()
            due = datetime.now() + timedelta(days=sla_days.get(sev, 7))
            ws.cell(row, 12, due.strftime('%Y-%m-%d'))
            ws.cell(row, 13, f"P{['critical','high','medium','low','info'].index(sev)+1} - {sev.capitalize()}" if sev in sla_days else 'P2 - High')
            ws.cell(row, 14, '')

            status_dv.add(ws.cell(row, 10))
            priority_dv.add(ws.cell(row, 13))

            # Color severity
            severity = finding.get('severity', '').lower()
            if severity in self.SEVERITY_COLORS:
                severity_cell = ws.cell(row, 3)
                severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                end_color=self.SEVERITY_COLORS[severity],
                                                fill_type="solid")
                if severity in ['critical', 'high']:
                    severity_cell.font = Font(color="FFFFFF", bold=True)

            # Color status
            status_val = finding.get('status', 'Open')
            status_colors = {'Open': 'FF0000', 'In Progress': 'FFA500', 'Resolved': '92D050', 'False Positive': '808080'}
            if status_val in status_colors:
                ws.cell(row, 10).fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
                if status_val == 'Open':
                    ws.cell(row, 10).font = Font(color="FFFFFF", bold=True)

            ws.row_dimensions[row].height = 50
            for c in range(1, 15):
                ws.cell(row, c).border = styles['border']
                ws.cell(row, c).alignment = styles['left_align']

            row += 1

        column_widths = [5, 30, 12, 8, 12, 40, 8, 40, 40, 15, 15, 12, 15, 30]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = 'A5'

    def _create_threat_model_sheet(self, wb: Workbook, threat_model: Dict[str, Any], styles: Dict):
        """Create detailed threat model (STRIDE) sheet"""
        ws = wb.create_sheet("Threat Model - STRIDE")

        # Title
        ws['A1'] = "Threat Model - STRIDE Analysis"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:H1')

        # Description
        row = 3
        ws[f'A{row}'] = "STRIDE Threat Modeling Framework"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        ws[f'A{row}'] = "Spoofing | Tampering | Repudiation | Information Disclosure | Denial of Service | Elevation of Privilege"
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        # STRIDE Categories
        stride_analysis = threat_model.get('stride_analysis', {})
        stride_categories = ['Spoofing', 'Tampering', 'Repudiation', 'Information Disclosure',
                            'Denial of Service', 'Elevation of Privilege']

        for category in stride_categories:
            threats = stride_analysis.get(category, stride_analysis.get(category.lower(), []))

            if threats:
                # Category Header
                ws[f'A{row}'] = f"{category} Threats ({len(threats)} identified)"
                ws[f'A{row}'].font = styles['subtitle_font']
                ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws.merge_cells(f'A{row}:H{row}')
                row += 1

                # Headers
                headers = ['#', 'Component', 'Threat', 'Description', 'Impact', 'Likelihood',
                          'Mitigation', 'Status']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row, col, header)
                    cell.fill = styles['header_fill']
                    cell.font = styles['header_font']

                row += 1

                # Threats
                for idx, threat in enumerate(threats, start=1):
                    ws.cell(row, 1, idx)
                    ws.cell(row, 2, threat.get('component', threat.get('asset', 'N/A')))
                    ws.cell(row, 3, threat.get('threat', threat.get('title', 'N/A')))
                    ws.cell(row, 4, threat.get('description', 'N/A'))
                    ws.cell(row, 5, threat.get('impact', 'High'))
                    ws.cell(row, 6, threat.get('likelihood', 'Medium'))
                    ws.cell(row, 7, threat.get('mitigation', threat.get('recommendation', 'N/A')))
                    ws.cell(row, 8, threat.get('status', 'Open'))

                    ws.row_dimensions[row].height = 50

                    for col in range(1, 9):
                        ws.cell(row, col).border = styles['border']
                        ws.cell(row, col).alignment = styles['left_align']

                    row += 1

                row += 2  # Space between categories

        # Trust Boundaries
        trust_boundaries = threat_model.get('trust_boundaries', [])
        if trust_boundaries:
            ws[f'A{row}'] = "Trust Boundaries Identified"
            ws[f'A{row}'].font = styles['subtitle_font']
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            for idx, boundary in enumerate(trust_boundaries, start=1):
                ws[f'A{row}'] = f"{idx}. {boundary}"
                ws.merge_cells(f'A{row}:H{row}')
                row += 1

        # Auto-size columns
        column_widths = [5, 20, 30, 35, 12, 12, 40, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_mitre_attack_sheet(self, wb: Workbook, threat_model: Dict[str, Any], styles: Dict):
        """Create MITRE ATT&CK mapping sheet"""
        ws = wb.create_sheet("MITRE ATT&CK Mapping")

        # Title
        ws['A1'] = "MITRE ATT&CK Technique Mapping"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:G1')

        # Description
        row = 3
        ws[f'A{row}'] = "Mapping of identified threats to MITRE ATT&CK framework"
        ws.merge_cells(f'A{row}:G{row}')
        row += 2

        # Headers
        headers = ['Technique ID', 'Technique Name', 'Tactic', 'Description',
                  'Related Threat', 'Mitigation', 'Detection']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        # Get MITRE mappings
        mitre_techniques = threat_model.get('mitre_attack_techniques', [])

        # If no direct MITRE mappings, extract from STRIDE analysis
        if not mitre_techniques:
            stride_analysis = threat_model.get('stride_analysis', {})
            for category, threats in stride_analysis.items():
                for threat in threats:
                    if 'mitre' in threat or 'technique' in threat:
                        mitre_techniques.append(threat)

        # Common MITRE ATT&CK techniques for application security
        default_techniques = [
            {
                'id': 'T1190',
                'name': 'Exploit Public-Facing Application',
                'tactic': 'Initial Access',
                'description': 'Adversaries may attempt to exploit weaknesses in Internet-facing applications',
                'mitigation': 'Application isolation, security scanning, input validation',
                'detection': 'Application logs, WAF alerts'
            },
            {
                'id': 'T1059',
                'name': 'Command and Scripting Interpreter',
                'tactic': 'Execution',
                'description': 'Adversaries may abuse command interpreters to execute commands',
                'mitigation': 'Input validation, least privilege, code review',
                'detection': 'Process monitoring, command-line analysis'
            },
            {
                'id': 'T1078',
                'name': 'Valid Accounts',
                'tactic': 'Persistence',
                'description': 'Adversaries may obtain credentials to gain access',
                'mitigation': 'Multi-factor authentication, credential rotation',
                'detection': 'Authentication logs, anomalous login detection'
            },
            {
                'id': 'T1548',
                'name': 'Abuse Elevation Control Mechanism',
                'tactic': 'Privilege Escalation',
                'description': 'Adversaries may circumvent access controls',
                'mitigation': 'Principle of least privilege, input validation',
                'detection': 'Process monitoring, API monitoring'
            },
            {
                'id': 'T1070',
                'name': 'Indicator Removal',
                'tactic': 'Defense Evasion',
                'description': 'Adversaries may delete or modify artifacts',
                'mitigation': 'Centralized logging, file integrity monitoring',
                'detection': 'Log analysis, file monitoring'
            },
            {
                'id': 'T1552',
                'name': 'Unsecured Credentials',
                'tactic': 'Credential Access',
                'description': 'Adversaries may search for credentials in insecure locations',
                'mitigation': 'Secrets management, encryption, code review',
                'detection': 'File monitoring, secret scanning'
            },
            {
                'id': 'T1087',
                'name': 'Account Discovery',
                'tactic': 'Discovery',
                'description': 'Adversaries may attempt to enumerate accounts',
                'mitigation': 'Rate limiting, authentication monitoring',
                'detection': 'API monitoring, authentication logs'
            },
            {
                'id': 'T1530',
                'name': 'Data from Cloud Storage',
                'tactic': 'Collection',
                'description': 'Adversaries may access data from cloud storage',
                'mitigation': 'Access controls, encryption, monitoring',
                'detection': 'Cloud audit logs, access monitoring'
            },
            {
                'id': 'T1041',
                'name': 'Exfiltration Over C2 Channel',
                'tactic': 'Exfiltration',
                'description': 'Adversaries may steal data over their command and control channel',
                'mitigation': 'Network segmentation, DLP, egress filtering',
                'detection': 'Network monitoring, anomaly detection'
            },
            {
                'id': 'T1499',
                'name': 'Endpoint Denial of Service',
                'tactic': 'Impact',
                'description': 'Adversaries may perform DoS attacks',
                'mitigation': 'Rate limiting, resource quotas, WAF',
                'detection': 'Performance monitoring, traffic analysis'
            },
        ]

        # Use provided techniques or defaults
        techniques_to_display = mitre_techniques if mitre_techniques else default_techniques

        for technique in techniques_to_display:
            ws.cell(row, 1, technique.get('id', technique.get('technique_id', 'N/A')))
            ws.cell(row, 2, technique.get('name', technique.get('technique_name', 'N/A')))
            ws.cell(row, 3, technique.get('tactic', 'N/A'))
            ws.cell(row, 4, technique.get('description', 'N/A'))
            ws.cell(row, 5, technique.get('related_threat', technique.get('threat', 'See threat model')))
            ws.cell(row, 6, technique.get('mitigation', 'N/A'))
            ws.cell(row, 7, technique.get('detection', 'N/A'))

            ws.row_dimensions[row].height = 50

            for col in range(1, 8):
                ws.cell(row, col).border = styles['border']
                ws.cell(row, col).alignment = styles['left_align']

            row += 1

        # Auto-size columns
        column_widths = [12, 30, 18, 40, 30, 40, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_github_monitor_sheet(self, wb: Workbook, github_data: Dict[str, Any], styles: Dict):
        """Create GitHub Commit Monitor summary sheet"""
        ws = wb.create_sheet("GitHub Monitor")

        ws['A1'] = "GitHub Commit Monitor — Security Analysis"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:G1')

        row = 3
        # Section 1: Monitored Repositories
        ws[f'A{row}'] = "Monitored Repositories"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

        repos = github_data.get('repos', [])
        if repos:
            headers = ['#', 'Repository', 'Branch', 'Commits Scanned', 'Last Scanned', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
            row += 1

            for idx, repo in enumerate(repos, 1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, repo.get('full_name', 'N/A'))
                ws.cell(row, 3, repo.get('default_branch', 'main'))
                ws.cell(row, 4, repo.get('total_commits_scanned', 0))
                ws.cell(row, 5, repo.get('last_scanned_at', 'N/A'))
                ws.cell(row, 6, 'Active' if repo.get('active', 1) else 'Inactive')
                for c in range(1, 7):
                    ws.cell(row, c).border = styles['border']
                row += 1
        else:
            ws[f'A{row}'] = "No repositories monitored"
            row += 1

        # Section 2: Risk Distribution
        row += 2
        ws[f'A{row}'] = "Commit Risk Distribution"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

        risk_dist = github_data.get('risk_distribution', {})
        total_commits = github_data.get('total_commits_scanned', 0)
        ws[f'A{row}'] = f"Total Commits Scanned: {total_commits}"
        ws[f'A{row}'].font = styles['bold_font']
        row += 1

        risk_headers = ['Risk Level', 'Count', 'Percentage']
        for col, header in enumerate(risk_headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
        row += 1

        risk_colors = {'critical': 'C00000', 'high': 'FF0000', 'medium': 'FFA500', 'low': 'FFFF00', 'clean': '92D050'}
        for level in ['critical', 'high', 'medium', 'low', 'clean']:
            count = risk_dist.get(level, 0)
            pct = (count / total_commits * 100) if total_commits > 0 else 0
            ws.cell(row, 1, level.upper())
            ws.cell(row, 2, count)
            ws.cell(row, 3, f"{pct:.1f}%")
            if level in risk_colors:
                ws.cell(row, 1).fill = PatternFill(start_color=risk_colors[level], end_color=risk_colors[level], fill_type="solid")
                if level in ['critical', 'high']:
                    ws.cell(row, 1).font = Font(color="FFFFFF", bold=True)
            for c in range(1, 4):
                ws.cell(row, c).border = styles['border']
            row += 1

        # Section 3: Developer Risk Profiles
        row += 2
        ws[f'A{row}'] = "Developer Risk Profiles"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

        devs = github_data.get('developer_risk_summary', [])
        if devs:
            dev_headers = ['#', 'Developer', 'Email', 'Total Commits', 'High Risk', 'Avg Risk Score', 'Trend']
            for col, header in enumerate(dev_headers, start=1):
                cell = ws.cell(row, col, header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
            row += 1

            for idx, dev in enumerate(devs, 1):
                ws.cell(row, 1, idx)
                ws.cell(row, 2, dev.get('author_name', 'Unknown'))
                ws.cell(row, 3, dev.get('author_email', 'N/A'))
                ws.cell(row, 4, dev.get('total_commits', 0))
                ws.cell(row, 5, dev.get('high_risk_commits', 0))
                ws.cell(row, 6, round(dev.get('avg_risk_score', 0), 2))
                ws.cell(row, 7, dev.get('risk_trend', 'stable'))
                for c in range(1, 8):
                    ws.cell(row, c).border = styles['border']
                row += 1

        # Auto-size
        column_widths = [5, 35, 15, 18, 15, 12, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_remediation_summary_sheet(self, wb: Workbook, scan_data: Dict[str, Any], styles: Dict):
        """Create consolidated remediation summary"""
        ws = wb.create_sheet("Remediation Summary")

        # Title
        ws['A1'] = "Remediation Action Plan"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:G1')

        row = 3
        ws[f'A{row}'] = "Prioritized Remediation Recommendations"
        ws[f'A{row}'].font = styles['subtitle_font']
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

        # Headers
        headers = ['Priority', 'Type', 'Issue', 'Location', 'Severity', 'Effort', 'Remediation Action']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']

        row += 1

        # Collect all remediations
        remediations = []

        # SAST
        for finding in scan_data.get('sast_findings', []):
            remediations.append({
                'type': 'SAST',
                'issue': finding.get('title', 'N/A'),
                'location': f"{finding.get('file_path', 'N/A')}:{finding.get('line_number', 'N/A')}",
                'severity': finding.get('severity', 'info'),
                'remediation': finding.get('remediation', 'N/A'),
                'cvss': float(finding.get('cvss_score', 0))
            })

        # SCA
        for finding in scan_data.get('sca_findings', []):
            remediations.append({
                'type': 'SCA',
                'issue': f"{finding.get('package', 'N/A')} - {finding.get('vulnerability', 'N/A')}",
                'location': f"{finding.get('package', 'N/A')} v{finding.get('installed_version', 'N/A')}",
                'severity': finding.get('severity', 'info'),
                'remediation': finding.get('remediation', f"Update to version {finding.get('fixed_version', 'latest')}"),
                'cvss': float(finding.get('cvss_score', 0))
            })

        # Secrets
        for finding in scan_data.get('secret_findings', []):
            remediations.append({
                'type': 'SECRET',
                'issue': finding.get('secret_type', 'Hardcoded Secret'),
                'location': f"{finding.get('file_path', 'N/A')}:{finding.get('line_number', 'N/A')}",
                'severity': finding.get('severity', 'high'),
                'remediation': finding.get('remediation', 'Remove hardcoded secret, use environment variables'),
                'cvss': 9.0 if finding.get('severity', '').lower() == 'critical' else 7.0
            })

        # Sort by severity and CVSS
        severity_priority = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}
        remediations.sort(key=lambda x: (severity_priority.get(x['severity'].lower(), 5), -x['cvss']))

        # Add to sheet
        for idx, rem in enumerate(remediations[:50], start=1):  # Top 50
            ws.cell(row, 1, idx)
            ws.cell(row, 2, rem['type'])
            ws.cell(row, 3, rem['issue'])
            ws.cell(row, 4, rem['location'])
            ws.cell(row, 5, rem['severity'].upper())

            # Estimate effort
            effort = 'Low' if rem['type'] == 'SCA' else ('High' if rem['severity'].lower() == 'critical' else 'Medium')
            ws.cell(row, 6, effort)
            ws.cell(row, 7, rem['remediation'])

            # Color code severity
            severity = rem['severity'].lower()
            if severity in self.SEVERITY_COLORS:
                severity_cell = ws.cell(row, 5)
                severity_cell.fill = PatternFill(start_color=self.SEVERITY_COLORS[severity],
                                                end_color=self.SEVERITY_COLORS[severity],
                                                fill_type="solid")
                if severity in ['critical', 'high']:
                    severity_cell.font = Font(color="FFFFFF", bold=True)

            ws.row_dimensions[row].height = 50

            for col in range(1, 8):
                ws.cell(row, col).border = styles['border']
                ws.cell(row, col).alignment = styles['left_align']

            row += 1

        # Auto-size columns
        column_widths = [8, 10, 35, 35, 12, 10, 50]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _count_by_severity(self, findings: List[Dict]) -> Dict[str, int]:
        """Count findings by severity"""
        counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'info': 0}
        for finding in findings:
            severity = finding.get('severity', 'info').lower()
            if severity in counts:
                counts[severity] += 1
        return counts

    def _count_all_severities(self, scan_data: Dict[str, Any]) -> Dict[str, int]:
        """Count all findings across all scan types by severity"""
        all_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'info': 0}

        for scan_type in ['sast_findings', 'sca_findings', 'secret_findings']:
            findings = scan_data.get(scan_type, [])
            counts = self._count_by_severity(findings)
            for severity, count in counts.items():
                all_counts[severity] += count

        return all_counts

    def _get_total_findings(self, scan_data: Dict[str, Any]) -> int:
        """Get total number of findings across all scan types"""
        total = 0
        total += len(scan_data.get('sast_findings', []))
        total += len(scan_data.get('sca_findings', []))
        total += len(scan_data.get('secret_findings', []))
        return total

    def _calculate_risk_score(self, severity_counts: Dict[str, int]) -> float:
        """Calculate risk score based on severity distribution"""
        weights = {'critical': 10, 'high': 7, 'medium': 4, 'low': 1, 'info': 0}
        score = sum(severity_counts.get(sev, 0) * weight for sev, weight in weights.items())
        return score

    def generate_pdf_report(self, scan_data: Dict[str, Any], output_path: str = None) -> BytesIO:
        """Generate enterprise-grade PDF report matching commercial tools (Checkmarx/Snyk/Veracode)"""
        from reportlab.graphics.shapes import Rect, String, Line

        buffer = BytesIO()

        # Custom styles
        styles = getSampleStyleSheet()
        brand_color = colors.HexColor('#1a237e')
        brand_accent = colors.HexColor('#0d47a1')
        brand_light = colors.HexColor('#e8eaf6')
        severity_colors_map = {
            'critical': colors.HexColor('#b71c1c'),
            'high': colors.HexColor('#d32f2f'),
            'medium': colors.HexColor('#f57c00'),
            'low': colors.HexColor('#fbc02d'),
            'info': colors.HexColor('#388e3c'),
        }

        title_style = ParagraphStyle('PDFTitle', parent=styles['Title'],
                                     fontSize=28, textColor=brand_color, spaceAfter=6,
                                     fontName='Helvetica-Bold')
        heading1_style = ParagraphStyle('H1', parent=styles['Heading1'],
                                        fontSize=18, textColor=brand_color, spaceAfter=12,
                                        spaceBefore=20, fontName='Helvetica-Bold',
                                        borderWidth=0, borderPadding=0,
                                        borderColor=brand_color)
        heading2_style = ParagraphStyle('H2', parent=styles['Heading2'],
                                        fontSize=14, textColor=brand_accent, spaceAfter=8,
                                        spaceBefore=12, fontName='Helvetica-Bold')
        heading3_style = ParagraphStyle('H3', parent=styles['Heading3'],
                                        fontSize=11, textColor=brand_accent, spaceAfter=6,
                                        spaceBefore=8, fontName='Helvetica-Bold')
        body_style = ParagraphStyle('Body', parent=styles['Normal'],
                                    fontSize=9, leading=13, spaceAfter=6,
                                    alignment=TA_JUSTIFY)
        small_style = ParagraphStyle('Small', parent=styles['Normal'],
                                     fontSize=8, leading=10, textColor=colors.HexColor('#555555'))
        code_style = ParagraphStyle('Code', parent=styles['Normal'],
                                    fontSize=7, leading=9, fontName='Courier',
                                    backColor=colors.HexColor('#f5f5f5'),
                                    borderWidth=0.5, borderColor=colors.HexColor('#cccccc'),
                                    borderPadding=4)

        # Page numbering
        page_num = [0]

        def header_footer(canvas, doc):
            page_num[0] += 1
            canvas.saveState()
            # Header line
            canvas.setStrokeColor(brand_color)
            canvas.setLineWidth(2)
            canvas.line(0.5*inch, letter[1] - 0.4*inch, letter[0] - 0.5*inch, letter[1] - 0.4*inch)
            # Header text
            canvas.setFont('Helvetica-Bold', 8)
            canvas.setFillColor(brand_color)
            canvas.drawString(0.75*inch, letter[1] - 0.35*inch, f"Security Assessment Report — {scan_data.get('project_name', 'Unknown')}")
            canvas.drawRightString(letter[0] - 0.75*inch, letter[1] - 0.35*inch, "CONFIDENTIAL")
            # Footer
            canvas.setStrokeColor(brand_color)
            canvas.setLineWidth(1)
            canvas.line(0.5*inch, 0.4*inch, letter[0] - 0.5*inch, 0.4*inch)
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#666666'))
            canvas.drawString(0.75*inch, 0.25*inch, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | AppSec Platform")
            canvas.drawRightString(letter[0] - 0.75*inch, 0.25*inch, f"Page {page_num[0]}")
            canvas.restoreState()

        def cover_page(canvas, doc):
            canvas.saveState()
            # Dark banner at top
            canvas.setFillColor(brand_color)
            canvas.rect(0, letter[1] - 3.5*inch, letter[0], 3.5*inch, fill=1)
            # Title text
            canvas.setFillColor(colors.white)
            canvas.setFont('Helvetica-Bold', 32)
            canvas.drawCentredString(letter[0]/2, letter[1] - 1.5*inch, "Application Security")
            canvas.setFont('Helvetica-Bold', 24)
            canvas.drawCentredString(letter[0]/2, letter[1] - 2.1*inch, "Assessment Report")
            canvas.setFont('Helvetica', 14)
            canvas.drawCentredString(letter[0]/2, letter[1] - 2.8*inch, scan_data.get('project_name', 'Unknown Project'))
            # Accent line
            canvas.setStrokeColor(colors.HexColor('#42a5f5'))
            canvas.setLineWidth(3)
            canvas.line(2*inch, letter[1] - 3.0*inch, letter[0] - 2*inch, letter[1] - 3.0*inch)
            # Footer
            canvas.setFillColor(colors.HexColor('#666666'))
            canvas.setFont('Helvetica', 9)
            canvas.drawCentredString(letter[0]/2, 0.5*inch, "CONFIDENTIAL — For Authorized Recipients Only")
            canvas.restoreState()

        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              topMargin=0.65*inch, bottomMargin=0.6*inch,
                              leftMargin=0.75*inch, rightMargin=0.75*inch)

        story = []

        # ═══════════════════════════════════════════════
        # COVER PAGE
        # ═══════════════════════════════════════════════
        story.append(Spacer(1, 3.5*inch))  # Space for the banner drawn by cover_page

        severity_counts = self._count_all_severities(scan_data)
        risk_score = self._calculate_risk_score(severity_counts)
        total_findings = self._get_total_findings(scan_data)
        sast_findings = scan_data.get('sast_findings', [])
        sca_findings = scan_data.get('sca_findings', [])
        secret_findings = scan_data.get('secret_findings', [])

        # Risk rating
        if risk_score >= 200:
            risk_rating = "CRITICAL"
            risk_color = '#b71c1c'
        elif risk_score >= 100:
            risk_rating = "HIGH"
            risk_color = '#d32f2f'
        elif risk_score >= 50:
            risk_rating = "MEDIUM"
            risk_color = '#f57c00'
        else:
            risk_rating = "LOW"
            risk_color = '#388e3c'

        # Metadata table on cover
        cover_meta = [
            ['Report Date', datetime.now().strftime("%B %d, %Y")],
            ['Project', scan_data.get('project_name', 'Unknown')],
            ['Industry', scan_data.get('industry_sector', 'Technology').title()],
            ['Technology Stack', ', '.join(scan_data.get('technology_stack', ['N/A'])) if scan_data.get('technology_stack') else 'N/A'],
            ['Scan Types', ', '.join(scan_data.get('scan_types', []))],
            ['Total Findings', str(total_findings)],
            ['Risk Rating', risk_rating],
            ['Risk Score', f"{risk_score:.0f}"],
        ]
        meta_table = Table(cover_meta, colWidths=[2*inch, 4*inch])
        meta_style_list = [
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), brand_color),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#333333')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e0e0e0')),
        ]
        # Color the risk rating row
        risk_row = 6
        meta_style_list.append(('TEXTCOLOR', (1, risk_row), (1, risk_row), colors.HexColor(risk_color)))
        meta_style_list.append(('FONTNAME', (1, risk_row), (1, risk_row), 'Helvetica-Bold'))
        meta_table.setStyle(TableStyle(meta_style_list))
        story.append(meta_table)
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # TABLE OF CONTENTS
        # ═══════════════════════════════════════════════
        story.append(Paragraph("Table of Contents", heading1_style))
        story.append(Spacer(1, 0.2*inch))
        toc_items = [
            "1. Executive Summary",
            "2. Risk Overview & Severity Distribution",
            "3. SAST — Static Application Security Testing",
            "4. SCA — Software Composition Analysis",
            "5. Secrets Detection",
            "6. Threat Model — STRIDE Analysis",
            "7. MITRE ATT&CK Mapping",
            "8. Compliance Mapping",
            "9. Detailed Critical & High Findings",
            "10. Remediation Roadmap",
            "Appendix A: Methodology & Scoring",
        ]
        for item in toc_items:
            story.append(Paragraph(item, ParagraphStyle('TOC', parent=body_style, fontSize=11, spaceAfter=8, leftIndent=20)))
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 1. EXECUTIVE SUMMARY
        # ═══════════════════════════════════════════════
        story.append(Paragraph("1. Executive Summary", heading1_style))

        exec_text = f"""A comprehensive application security assessment was conducted on <b>{scan_data.get('project_name', 'the application')}</b>
        covering static code analysis (SAST), software composition analysis (SCA), and secrets detection.
        The assessment identified <b>{total_findings} security findings</b> across the application with an aggregate
        risk score of <b>{risk_score:.0f}</b> (rated <font color="{risk_color}"><b>{risk_rating}</b></font>)."""
        story.append(Paragraph(exec_text, body_style))
        story.append(Spacer(1, 0.15*inch))

        # Key metrics boxes
        metrics_data = [
            ['Total\nFindings', 'Critical', 'High', 'Medium', 'Low', 'Risk\nScore'],
            [str(total_findings), str(severity_counts.get('critical', 0)),
             str(severity_counts.get('high', 0)), str(severity_counts.get('medium', 0)),
             str(severity_counts.get('low', 0)), f"{risk_score:.0f}"]
        ]
        metrics_table = Table(metrics_data, colWidths=[1.05*inch]*6)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, 1), 20),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 1), (0, 1), brand_color),
            ('TEXTCOLOR', (1, 1), (1, 1), severity_colors_map['critical']),
            ('TEXTCOLOR', (2, 1), (2, 1), severity_colors_map['high']),
            ('TEXTCOLOR', (3, 1), (3, 1), severity_colors_map['medium']),
            ('TEXTCOLOR', (4, 1), (4, 1), severity_colors_map['low']),
            ('TEXTCOLOR', (5, 1), (5, 1), colors.HexColor(risk_color)),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.2*inch))

        # Findings breakdown
        breakdown_data = [['Scan Type', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Total']]
        for scan_label, findings_list in [('SAST', sast_findings), ('SCA', sca_findings), ('Secrets', secret_findings)]:
            sev = self._count_by_severity(findings_list)
            breakdown_data.append([scan_label, str(sev['critical']), str(sev['high']),
                                   str(sev['medium']), str(sev['low']), str(sev['info']),
                                   str(len(findings_list))])
        breakdown_data.append(['TOTAL', str(severity_counts['critical']), str(severity_counts['high']),
                              str(severity_counts['medium']), str(severity_counts['low']),
                              str(severity_counts['info']), str(total_findings)])

        breakdown_table = Table(breakdown_data, colWidths=[1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        breakdown_style = [
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), brand_light),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]
        breakdown_table.setStyle(TableStyle(breakdown_style))
        story.append(breakdown_table)
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 2. RISK OVERVIEW WITH CHARTS
        # ═══════════════════════════════════════════════
        story.append(Paragraph("2. Risk Overview & Severity Distribution", heading1_style))

        # Severity pie chart
        story.append(Paragraph("Severity Distribution", heading2_style))
        d = Drawing(400, 200)
        pie = Pie()
        pie.x = 100
        pie.y = 20
        pie.width = 150
        pie.height = 150
        pie_data = []
        pie_labels = []
        pie_colors = []
        for sev in self.SEVERITY_ORDER:
            count = severity_counts.get(sev, 0)
            if count > 0:
                pie_data.append(count)
                pie_labels.append(f"{sev.upper()}: {count}")
                pie_colors.append(severity_colors_map[sev])
        if pie_data:
            pie.data = pie_data
            pie.labels = pie_labels
            for i, c in enumerate(pie_colors):
                pie.slices[i].fillColor = c
                pie.slices[i].strokeColor = colors.white
                pie.slices[i].strokeWidth = 1
            d.add(pie)
            story.append(d)
        story.append(Spacer(1, 0.2*inch))

        # Scan type bar chart
        story.append(Paragraph("Findings by Scan Type", heading2_style))
        d2 = Drawing(450, 180)
        bc = VerticalBarChart()
        bc.x = 60
        bc.y = 20
        bc.height = 130
        bc.width = 350
        bc.data = [
            [self._count_by_severity(sast_findings).get(s, 0) for s in ['critical', 'high', 'medium', 'low']],
            [self._count_by_severity(sca_findings).get(s, 0) for s in ['critical', 'high', 'medium', 'low']],
            [self._count_by_severity(secret_findings).get(s, 0) for s in ['critical', 'high', 'medium', 'low']],
        ]
        bc.categoryAxis.categoryNames = ['Critical', 'High', 'Medium', 'Low']
        bc.categoryAxis.labels.fontSize = 8
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 8
        bar_colors = [brand_color, colors.HexColor('#42a5f5'), colors.HexColor('#90caf9')]
        for i, c in enumerate(bar_colors):
            bc.bars[i].fillColor = c
        d2.add(bc)
        # Legend
        legend_y = 160
        for i, label in enumerate(['SAST', 'SCA', 'Secrets']):
            d2.add(Rect(370, legend_y - i*15, 10, 10, fillColor=bar_colors[i], strokeColor=None))
            d2.add(String(385, legend_y - i*15 + 1, label, fontSize=8))
        story.append(d2)
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 3. SAST FINDINGS
        # ═══════════════════════════════════════════════
        if sast_findings:
            story.append(Paragraph("3. SAST — Static Application Security Testing", heading1_style))

            sast_severity = self._count_by_severity(sast_findings)
            story.append(Paragraph(
                f"Total SAST findings: <b>{len(sast_findings)}</b> | "
                f"Critical: <font color='#b71c1c'><b>{sast_severity['critical']}</b></font> | "
                f"High: <font color='#d32f2f'><b>{sast_severity['high']}</b></font> | "
                f"Medium: <font color='#f57c00'><b>{sast_severity['medium']}</b></font> | "
                f"Low: <font color='#fbc02d'><b>{sast_severity['low']}</b></font>",
                body_style))
            story.append(Spacer(1, 0.1*inch))

            # Top CWE categories
            cwe_groups = defaultdict(int)
            for f in sast_findings:
                cwe = f.get('cwe_id', 'Unknown')
                cwe_groups[cwe] += 1
            top_cwes = sorted(cwe_groups.items(), key=lambda x: x[1], reverse=True)[:10]
            if top_cwes:
                story.append(Paragraph("Top CWE Categories", heading3_style))
                cwe_data = [['CWE ID', 'Count', 'Percentage']]
                for cwe_id, count in top_cwes:
                    pct = (count / len(sast_findings) * 100) if sast_findings else 0
                    cwe_data.append([cwe_id, str(count), f"{pct:.1f}%"])
                cwe_table = Table(cwe_data, colWidths=[2*inch, 1*inch, 1.5*inch])
                cwe_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), brand_accent),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(cwe_table)
                story.append(Spacer(1, 0.15*inch))

            # All SAST findings table
            story.append(Paragraph("SAST Findings Summary", heading3_style))
            sast_sorted = sorted(sast_findings, key=lambda x: (['critical','high','medium','low','info'].index(x.get('severity','info').lower()), -float(x.get('cvss_score', 0) or 0)))

            sast_table_data = [['#', 'Severity', 'Title', 'CWE', 'OWASP', 'File', 'CVSS', 'Status']]
            for idx, f in enumerate(sast_sorted, 1):
                title_text = (f.get('title', 'N/A') or 'N/A')[:45]
                file_text = (f.get('file_path', 'N/A') or 'N/A')
                if len(file_text) > 30:
                    file_text = '...' + file_text[-27:]
                sast_table_data.append([
                    str(idx), (f.get('severity', '') or '').upper(),
                    title_text, f.get('cwe_id', 'N/A') or 'N/A',
                    (f.get('owasp_category', 'N/A') or 'N/A')[:20],
                    f"{file_text}:{f.get('line_number', '')}",
                    str(f.get('cvss_score', 'N/A') or 'N/A'),
                    f.get('status', 'Open')
                ])

            sast_t = Table(sast_table_data, colWidths=[0.3*inch, 0.6*inch, 1.8*inch, 0.6*inch, 0.8*inch, 1.4*inch, 0.4*inch, 0.6*inch])
            sast_style = [
                ('BACKGROUND', (0, 0), (-1, 0), brand_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (6, 0), (6, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]
            # Color severity cells
            for i, f in enumerate(sast_sorted, 1):
                sev = (f.get('severity', 'info') or 'info').lower()
                if sev in severity_colors_map:
                    sast_style.append(('TEXTCOLOR', (1, i), (1, i), severity_colors_map[sev]))
                    sast_style.append(('FONTNAME', (1, i), (1, i), 'Helvetica-Bold'))
            sast_t.setStyle(TableStyle(sast_style))
            story.append(sast_t)
            story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 4. SCA FINDINGS
        # ═══════════════════════════════════════════════
        if sca_findings:
            story.append(Paragraph("4. SCA — Software Composition Analysis", heading1_style))

            sca_severity = self._count_by_severity(sca_findings)
            story.append(Paragraph(
                f"Total SCA findings: <b>{len(sca_findings)}</b> | "
                f"Critical: <font color='#b71c1c'><b>{sca_severity['critical']}</b></font> | "
                f"High: <font color='#d32f2f'><b>{sca_severity['high']}</b></font> | "
                f"Medium: <font color='#f57c00'><b>{sca_severity['medium']}</b></font> | "
                f"Low: <font color='#fbc02d'><b>{sca_severity['low']}</b></font>",
                body_style))
            story.append(Spacer(1, 0.1*inch))

            sca_sorted = sorted(sca_findings, key=lambda x: (['critical','high','medium','low','info'].index(x.get('severity','info').lower()), -float(x.get('cvss_score', 0) or 0)))

            sca_table_data = [['#', 'Severity', 'Package', 'CVE', 'CVSS', 'Remediation', 'Status']]
            for idx, f in enumerate(sca_sorted, 1):
                pkg = (f.get('package', 'N/A') or 'N/A')
                if len(pkg) > 35:
                    pkg = pkg[:32] + '...'
                rem = (f.get('remediation', 'N/A') or 'N/A')[:50]
                sca_table_data.append([
                    str(idx), (f.get('severity', '') or '').upper(),
                    pkg, f.get('cve', 'N/A') or 'N/A',
                    str(f.get('cvss_score', 'N/A') or 'N/A'),
                    rem, f.get('status', 'Open')
                ])

            sca_t = Table(sca_table_data, colWidths=[0.3*inch, 0.6*inch, 1.6*inch, 1.2*inch, 0.4*inch, 1.8*inch, 0.6*inch])
            sca_style = [
                ('BACKGROUND', (0, 0), (-1, 0), brand_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (4, 0), (4, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]
            for i, f in enumerate(sca_sorted, 1):
                sev = (f.get('severity', 'info') or 'info').lower()
                if sev in severity_colors_map:
                    sca_style.append(('TEXTCOLOR', (1, i), (1, i), severity_colors_map[sev]))
                    sca_style.append(('FONTNAME', (1, i), (1, i), 'Helvetica-Bold'))
            sca_t.setStyle(TableStyle(sca_style))
            story.append(sca_t)
            story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 5. SECRETS DETECTION
        # ═══════════════════════════════════════════════
        if secret_findings:
            story.append(Paragraph("5. Secrets Detection", heading1_style))
            story.append(Paragraph(
                f"<font color='#b71c1c'><b>{len(secret_findings)} hardcoded secrets</b></font> detected in source code. "
                "These pose an immediate risk of credential exposure and unauthorized access.", body_style))
            story.append(Spacer(1, 0.1*inch))

            sec_sorted = sorted(secret_findings, key=lambda x: (['critical','high','medium','low','info'].index(x.get('severity','info').lower())))
            sec_data = [['#', 'Severity', 'Secret Type', 'File Path', 'CVSS', 'Status']]
            for idx, f in enumerate(sec_sorted, 1):
                sec_data.append([
                    str(idx), (f.get('severity', '') or '').upper(),
                    (f.get('secret_type', f.get('title', 'Unknown')) or 'Unknown')[:35],
                    (f.get('file_path', 'N/A') or 'N/A')[:40],
                    str(f.get('cvss_score', 'N/A') or 'N/A'),
                    f.get('status', 'Open')
                ])

            sec_t = Table(sec_data, colWidths=[0.3*inch, 0.6*inch, 1.8*inch, 2.2*inch, 0.5*inch, 0.6*inch])
            sec_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#b71c1c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff3f3')]),
            ]
            sec_t.setStyle(TableStyle(sec_style))
            story.append(sec_t)
            story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 6. THREAT MODEL - STRIDE
        # ═══════════════════════════════════════════════
        threat_model = scan_data.get('threat_model', {})
        if threat_model:
            story.append(Paragraph("6. Threat Model — STRIDE Analysis", heading1_style))
            stride_analysis = threat_model.get('stride_analysis', {})
            stride_categories = ['Spoofing', 'Tampering', 'Repudiation', 'Information Disclosure',
                                'Denial of Service', 'Elevation of Privilege']

            stride_summary_data = [['STRIDE Category', 'Threats Identified', 'Risk Level']]
            for cat in stride_categories:
                threats = stride_analysis.get(cat, stride_analysis.get(cat.lower(), []))
                count = len(threats) if isinstance(threats, list) else 0
                risk = 'High' if count >= 3 else ('Medium' if count >= 1 else 'Low')
                stride_summary_data.append([cat, str(count), risk])

            stride_t = Table(stride_summary_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            stride_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), brand_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, brand_light]),
            ]))
            story.append(stride_t)
            story.append(Spacer(1, 0.15*inch))

            # Detail for each category with threats
            for cat in stride_categories:
                threats = stride_analysis.get(cat, stride_analysis.get(cat.lower(), []))
                if isinstance(threats, list) and threats:
                    story.append(Paragraph(f"{cat} ({len(threats)} threats)", heading3_style))
                    for t in threats[:5]:
                        threat_text = f"<b>{t.get('threat', t.get('title', 'N/A'))}</b>: {t.get('description', 'N/A')[:150]}"
                        story.append(Paragraph(threat_text, small_style))
                    story.append(Spacer(1, 0.05*inch))

            story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 7. MITRE ATT&CK
        # ═══════════════════════════════════════════════
        if threat_model:
            story.append(Paragraph("7. MITRE ATT&CK Mapping", heading1_style))
            mitre = threat_model.get('mitre_mapping', {})
            techniques_raw = mitre.get('techniques', {}) if isinstance(mitre, dict) else {}
            # Normalize: could be dict keyed by technique ID or a list
            if isinstance(techniques_raw, dict):
                techniques = list(techniques_raw.values())
            elif isinstance(techniques_raw, list):
                techniques = techniques_raw
            else:
                techniques = []

            if techniques:
                mitre_data = [['Technique ID', 'Name', 'Tactic', 'Threats', 'Severity']]
                for t in techniques[:15]:
                    mitre_data.append([
                        t.get('id', t.get('technique_id', 'N/A')),
                        (t.get('name', t.get('technique_name', 'N/A')) or 'N/A')[:30],
                        (t.get('tactic', 'N/A') or 'N/A')[:20],
                        str(t.get('threat_count', len(t.get('related_threats', [])))),
                        (t.get('max_severity', 'N/A') or 'N/A').upper()
                    ])
                mitre_t = Table(mitre_data, colWidths=[0.8*inch, 1.8*inch, 1.2*inch, 0.6*inch, 0.8*inch])
                mitre_t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), brand_color),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ]))
                story.append(mitre_t)
            else:
                story.append(Paragraph("MITRE ATT&CK mapping data not available for this project.", body_style))
            story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 8. COMPLIANCE MAPPING
        # ═══════════════════════════════════════════════
        story.append(Paragraph("8. Compliance Mapping", heading1_style))
        story.append(Paragraph(
            "The findings in this report map to the following industry compliance frameworks:", body_style))
        story.append(Spacer(1, 0.1*inch))

        compliance_data = [['Framework', 'Relevant Findings', 'Coverage']]
        owasp_count = sum(1 for f in sast_findings if f.get('owasp_category'))
        cwe_count = sum(1 for f in sast_findings if f.get('cwe_id'))
        compliance_data.append(['OWASP Top 10 (2021)', str(owasp_count), f"{(owasp_count/len(sast_findings)*100) if sast_findings else 0:.0f}% mapped"])
        compliance_data.append(['MITRE CWE', str(cwe_count), f"{(cwe_count/len(sast_findings)*100) if sast_findings else 0:.0f}% mapped"])
        compliance_data.append(['NIST SP 800-53', str(total_findings), 'SI-10, SA-11, RA-5'])
        compliance_data.append(['PCI DSS 4.0', str(severity_counts.get('critical', 0) + severity_counts.get('high', 0)), 'Req 6.2, 6.3, 6.5'])
        compliance_data.append(['SOC 2 Type II', str(total_findings), 'CC6.1, CC7.1, CC8.1'])

        comp_t = Table(compliance_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        comp_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, brand_light]),
        ]))
        story.append(comp_t)
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 9. DETAILED CRITICAL & HIGH FINDINGS
        # ═══════════════════════════════════════════════
        story.append(Paragraph("9. Detailed Critical & High Findings", heading1_style))
        story.append(Paragraph(
            "This section provides detailed information for all critical and high severity findings requiring immediate attention.",
            body_style))
        story.append(Spacer(1, 0.15*inch))

        all_critical_high = []
        for f in sast_findings:
            if (f.get('severity', '') or '').lower() in ['critical', 'high']:
                all_critical_high.append({**f, '_type': 'SAST'})
        for f in sca_findings:
            if (f.get('severity', '') or '').lower() in ['critical', 'high']:
                all_critical_high.append({**f, '_type': 'SCA'})
        for f in secret_findings:
            if (f.get('severity', '') or '').lower() in ['critical', 'high']:
                all_critical_high.append({**f, '_type': 'SECRET'})

        all_critical_high.sort(key=lambda x: (0 if (x.get('severity','').lower() == 'critical') else 1, -float(x.get('cvss_score', 0) or 0)))

        for idx, f in enumerate(all_critical_high[:25], 1):
            sev = (f.get('severity', '') or 'high').upper()
            sev_lower = sev.lower()
            sev_color = '#b71c1c' if sev_lower == 'critical' else '#d32f2f'

            finding_title = f"<font color='{sev_color}'>[{sev}]</font> {f.get('title', f.get('vulnerability', f.get('secret_type', 'N/A')))}"
            story.append(Paragraph(f"Finding #{idx}: {finding_title}", heading3_style))

            detail_rows = [
                ['Type', f['_type']],
                ['Severity', sev],
                ['CVSS Score', str(f.get('cvss_score', 'N/A') or 'N/A')],
            ]
            if f['_type'] == 'SAST':
                detail_rows.extend([
                    ['CWE', f.get('cwe_id', 'N/A') or 'N/A'],
                    ['OWASP', f.get('owasp_category', 'N/A') or 'N/A'],
                    ['File', f"{f.get('file_path', 'N/A')}:{f.get('line_number', '')}"],
                    ['STRIDE', f.get('stride_category', 'N/A') or 'N/A'],
                    ['MITRE ATT&CK', f"{f.get('mitre_attack_id', '')} {f.get('mitre_attack_name', '')}".strip() or 'N/A'],
                ])
            elif f['_type'] == 'SCA':
                detail_rows.extend([
                    ['Package', f.get('package', 'N/A') or 'N/A'],
                    ['CVE', f.get('cve', 'N/A') or 'N/A'],
                    ['CWE', f.get('cwe_id', 'N/A') or 'N/A'],
                ])
            elif f['_type'] == 'SECRET':
                detail_rows.extend([
                    ['File', f"{f.get('file_path', 'N/A')}:{f.get('line_number', '')}"],
                ])
            detail_rows.append(['Status', f.get('status', 'Open')])

            detail_t = Table(detail_rows, colWidths=[1.2*inch, 5*inch])
            detail_t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (0, -1), brand_color),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('LINEBELOW', (0, 0), (-1, -2), 0.25, colors.HexColor('#eeeeee')),
            ]))
            story.append(detail_t)

            # Description
            desc = f.get('description', '')
            if desc:
                story.append(Spacer(1, 0.05*inch))
                story.append(Paragraph(f"<b>Description:</b> {str(desc)[:300]}", small_style))

            # Business Impact
            bi = f.get('business_impact', '')
            if bi:
                story.append(Paragraph(f"<b>Business Impact:</b> {str(bi)[:200]}", small_style))

            # Remediation
            rem = f.get('remediation', '')
            if rem:
                story.append(Paragraph(f"<b>Remediation:</b> {str(rem)[:300]}", small_style))

            # Code snippet
            snippet = f.get('code_snippet', '')
            if snippet and f['_type'] in ['SAST', 'SECRET']:
                story.append(Spacer(1, 0.03*inch))
                snippet_text = str(snippet)[:200].replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br/>')
                story.append(Paragraph(snippet_text, code_style))

            story.append(Spacer(1, 0.15*inch))

            # Page break every 4 findings
            if idx % 4 == 0 and idx < len(all_critical_high[:25]):
                story.append(PageBreak())

        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # 10. REMEDIATION ROADMAP
        # ═══════════════════════════════════════════════
        story.append(Paragraph("10. Remediation Roadmap", heading1_style))

        story.append(Paragraph("<b>Phase 1: Immediate (0-7 days)</b>", heading2_style))
        phase1_items = [
            "Rotate all exposed secrets and credentials immediately",
            f"Address {severity_counts.get('critical', 0)} critical severity findings",
            "Patch critical CVEs in dependencies (Log4Shell, Spring4Shell, etc.)",
            "Implement input validation for injection vulnerabilities",
        ]
        for item in phase1_items:
            story.append(Paragraph(f"  \u2022  {item}", body_style))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph("<b>Phase 2: Short-term (7-30 days)</b>", heading2_style))
        phase2_items = [
            f"Remediate {severity_counts.get('high', 0)} high severity findings",
            "Update all vulnerable dependencies to patched versions",
            "Implement secure coding guidelines based on CWE findings",
            "Enable security scanning in CI/CD pipeline",
            "Conduct threat model review based on STRIDE analysis",
        ]
        for item in phase2_items:
            story.append(Paragraph(f"  \u2022  {item}", body_style))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph("<b>Phase 3: Medium-term (30-90 days)</b>", heading2_style))
        phase3_items = [
            f"Address {severity_counts.get('medium', 0)} medium severity findings",
            "Implement MITRE ATT&CK-based detection controls",
            "Deploy runtime application security protection (RASP)",
            "Establish vulnerability management SLA process",
            "Conduct developer security training based on finding categories",
        ]
        for item in phase3_items:
            story.append(Paragraph(f"  \u2022  {item}", body_style))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph("<b>Phase 4: Long-term (90+ days)</b>", heading2_style))
        phase4_items = [
            "Implement DevSecOps maturity program",
            "Establish regular penetration testing cadence",
            "Deploy Software Bill of Materials (SBOM) management",
            "Implement zero-trust architecture recommendations",
        ]
        for item in phase4_items:
            story.append(Paragraph(f"  \u2022  {item}", body_style))
        story.append(PageBreak())

        # ═══════════════════════════════════════════════
        # APPENDIX A: METHODOLOGY
        # ═══════════════════════════════════════════════
        story.append(Paragraph("Appendix A: Methodology & Scoring", heading1_style))

        story.append(Paragraph("<b>Severity Classification</b>", heading3_style))
        sev_desc = [
            ['Severity', 'CVSS Range', 'SLA', 'Description'],
            ['CRITICAL', '9.0 - 10.0', '7 days', 'Exploitable remotely with no authentication, leads to full system compromise'],
            ['HIGH', '7.0 - 8.9', '14 days', 'Significant impact requiring prompt remediation, may require specific conditions'],
            ['MEDIUM', '4.0 - 6.9', '30 days', 'Moderate impact, requires specific conditions or authenticated access'],
            ['LOW', '0.1 - 3.9', '90 days', 'Minor impact, defense-in-depth issue, limited exploitability'],
            ['INFO', '0.0', '180 days', 'Informational finding, best practice recommendation'],
        ]
        sev_t = Table(sev_desc, colWidths=[0.8*inch, 0.8*inch, 0.7*inch, 3.7*inch])
        sev_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(sev_t)
        story.append(Spacer(1, 0.15*inch))

        story.append(Paragraph("<b>Risk Score Calculation</b>", heading3_style))
        story.append(Paragraph(
            "Risk Score = (Critical × 10) + (High × 7) + (Medium × 4) + (Low × 1). "
            "Ratings: Critical ≥ 200, High ≥ 100, Medium ≥ 50, Low < 50.", body_style))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph("<b>Scanning Methodology</b>", heading3_style))
        methodology_text = """
        <b>SAST:</b> Static analysis of source code to identify security vulnerabilities including injection flaws,
        authentication issues, cryptographic weaknesses, and insecure coding patterns. Findings are mapped to CWE and OWASP Top 10.<br/><br/>
        <b>SCA:</b> Analysis of third-party dependencies and libraries against known vulnerability databases (NVD, GitHub Advisory).
        Includes version analysis, license compliance, and reachability assessment.<br/><br/>
        <b>Secrets Detection:</b> Pattern-based and entropy-based detection of hardcoded credentials, API keys, tokens,
        and other sensitive data in source code and configuration files.<br/><br/>
        <b>Threat Modeling:</b> STRIDE-based analysis of application architecture to identify potential threats and attack vectors,
        mapped to MITRE ATT&CK framework for comprehensive coverage.
        """
        story.append(Paragraph(methodology_text, body_style))

        # Build with custom page handling
        doc.build(story, onFirstPage=cover_page, onLaterPages=header_footer)
        buffer.seek(0)

        if output_path:
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())
            return output_path

        return buffer

    def generate_xml_report(self, scan_data: Dict[str, Any], output_path: str = None) -> str:
        """Generate detailed XML report for tool integration"""
        root = ET.Element('SecurityScanReport', version='2.0')

        # Metadata
        metadata = ET.SubElement(root, 'Metadata')
        ET.SubElement(metadata, 'ReportDate').text = datetime.now().isoformat()
        ET.SubElement(metadata, 'ProjectName').text = scan_data.get('project_name', 'Unknown')
        ET.SubElement(metadata, 'ScanTypes').text = ','.join(scan_data.get('scan_types', []))
        ET.SubElement(metadata, 'TotalFindings').text = str(self._get_total_findings(scan_data))

        # Risk Score
        severity_counts = self._count_all_severities(scan_data)
        risk_score = self._calculate_risk_score(severity_counts)
        ET.SubElement(metadata, 'RiskScore').text = str(risk_score)

        # Severity Summary
        severity_summary = ET.SubElement(root, 'SeveritySummary')
        for severity in self.SEVERITY_ORDER:
            count = severity_counts.get(severity, 0)
            ET.SubElement(severity_summary, severity.capitalize(), count=str(count))

        # SAST Findings
        sast_findings = scan_data.get('sast_findings', [])
        if sast_findings:
            sast = ET.SubElement(root, 'SASTFindings', count=str(len(sast_findings)))
            for finding in sast_findings:
                vuln = ET.SubElement(sast, 'Vulnerability',
                                    severity=finding.get('severity', ''),
                                    cvss=str(finding.get('cvss_score', '')))
                ET.SubElement(vuln, 'Title').text = finding.get('title', '')
                ET.SubElement(vuln, 'CWE').text = finding.get('cwe_id', '')
                ET.SubElement(vuln, 'CWEName').text = finding.get('cwe_name', '')
                ET.SubElement(vuln, 'OWASP').text = finding.get('owasp_category', '')
                ET.SubElement(vuln, 'FilePath').text = finding.get('file_path', '')
                ET.SubElement(vuln, 'LineNumber').text = str(finding.get('line_number', ''))
                ET.SubElement(vuln, 'Description').text = finding.get('description', '')
                ET.SubElement(vuln, 'Remediation').text = finding.get('remediation', '')
                ET.SubElement(vuln, 'CodeSnippet').text = finding.get('code_snippet', finding.get('vulnerable_code', ''))

        # SCA Findings
        sca_findings = scan_data.get('sca_findings', [])
        if sca_findings:
            sca = ET.SubElement(root, 'SCAFindings', count=str(len(sca_findings)))
            for finding in sca_findings:
                dep = ET.SubElement(sca, 'Dependency',
                                   severity=finding.get('severity', ''),
                                   cvss=str(finding.get('cvss_score', '')))
                ET.SubElement(dep, 'Package').text = finding.get('package', '')
                ET.SubElement(dep, 'InstalledVersion').text = finding.get('installed_version', '')
                ET.SubElement(dep, 'FixedVersion').text = finding.get('fixed_version', finding.get('safe_version', ''))
                ET.SubElement(dep, 'Vulnerability').text = finding.get('vulnerability', '')
                ET.SubElement(dep, 'CVE').text = finding.get('cve', '')
                ET.SubElement(dep, 'Description').text = finding.get('description', '')
                ET.SubElement(dep, 'Remediation').text = finding.get('remediation', '')
                ET.SubElement(dep, 'References').text = finding.get('references', '')

        # Secret Findings
        secret_findings = scan_data.get('secret_findings', [])
        if secret_findings:
            secrets = ET.SubElement(root, 'SecretFindings', count=str(len(secret_findings)))
            for finding in secret_findings:
                secret = ET.SubElement(secrets, 'Secret', severity=finding.get('severity', ''))
                ET.SubElement(secret, 'Type').text = finding.get('secret_type', '')
                ET.SubElement(secret, 'FilePath').text = finding.get('file_path', '')
                ET.SubElement(secret, 'LineNumber').text = str(finding.get('line_number', ''))
                ET.SubElement(secret, 'MaskedValue').text = finding.get('masked_value', '')
                ET.SubElement(secret, 'Remediation').text = finding.get('remediation', '')

        # Threat Model
        threat_model = scan_data.get('threat_model', {})
        if threat_model:
            tm = ET.SubElement(root, 'ThreatModel')
            stride_analysis = threat_model.get('stride_analysis', {})
            for category, threats in stride_analysis.items():
                cat_elem = ET.SubElement(tm, 'STRIDECategory', name=category)
                for threat in threats:
                    threat_elem = ET.SubElement(cat_elem, 'Threat')
                    ET.SubElement(threat_elem, 'Component').text = threat.get('component', threat.get('asset', ''))
                    ET.SubElement(threat_elem, 'Description').text = threat.get('description', '')
                    ET.SubElement(threat_elem, 'Mitigation').text = threat.get('mitigation', threat.get('recommendation', ''))

        # Pretty print XML
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(xml_str)
            return output_path

        return xml_str
